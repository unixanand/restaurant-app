import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side
import string
col_list = list(string.ascii_uppercase)



def xcel(rec, title,header,fields,fields_len, position) :
    data = rec
    
    
    df = pd.DataFrame(data)

    sub_col_list = col_list[position -1 : fields_len + (position -1)]

    # Create a new Excel workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = title

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Write the header
    header = header
    c = sub_col_list[0]
    c += str(1)
    ws[c] = header
    ws[c].alignment = Alignment(horizontal='center')

    rec = 1 
    for i in range(fields_len) :
        if i > 0 :
            continue
        AC = sub_col_list[i]+str(rec)
    AC += ':' + sub_col_list[i]+str(rec)
    ws.merge_cells(AC)
    
    columns = []
    column = 1
    for i in range(fields_len) :
        c = sub_col_list[i]+str(column)
        columns.append(c)
    for col in columns :
        ws[col].border = thin_border
        
    
    rec += 2
    # Write column headers
    for i in range(fields_len) :
        j = i #+ (position -1)
        c = sub_col_list[j]+str(rec)
        ws[c] = fields[i]

    for i in range(fields_len) :
        j = i #+ (position -1)
        c = sub_col_list[j]+str(rec)
        ws[c].alignment = Alignment(horizontal='center')
        ws[c].border = thin_border
    
    #for col in ['A3', 'B3', 'C3']:
        #ws[col].alignment = Alignment(horizontal='center')
        #ws[col].border = thin_border
    
    i = 1

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 4):
        for c_idx, value in enumerate(row, position):
            ws.cell(row=r_idx, column=c_idx, value=value)
            if i == 1 :
                ws.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='center')
            else :
                ws.cell(row=r_idx, column=c_idx).alignment = Alignment(horizontal='left')
            i += 1
        i = 1
        
    # Adjust column widths for better readability (matching your previous formatting)
    
    for i in range(fields_len) :
        j = i #+ (position -1)
        c = sub_col_list[j]
        ws.column_dimensions[c].width = 18 
        
    c = sub_col_list[0]+str(3)+':'+ sub_col_list[-1]
    
    for row in ws[c + str(ws.max_row)]:
        for cell in row:
            cell.border = thin_border
    from datetime import datetime
    curr_date = datetime.now()
    ts = curr_date.strftime("%Y%m%d%I%M%S")

    # Save the workbook
    print("Excel file is generated!")
    file_path = f"./reports/{title}_{ts}.xlsx"
    print("File path : ",f'{file_path}')
    wb.save(file_path)
